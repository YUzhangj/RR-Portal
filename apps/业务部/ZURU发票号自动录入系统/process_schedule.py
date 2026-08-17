#!/usr/bin/env python3
"""排期处理：识别总排期/分排期，逐行分类，可自动填的写入发票号（粉色，不覆盖已有值）"""
import openpyxl, re, os, json, datetime
from openpyxl.styles import PatternFill
from collections import Counter, defaultdict
from matcher_unified import UnifiedMatcher, norm_code, parse_date, code_candidates, line_qtys_for

PINK_FILL = PatternFill("solid", fgColor="FFB6C1")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")  # 不确定行标绿，供人工定夺

DATE_SUFFIX_RE = re.compile(r'-\d{1,2}\.\d{1,2}$')
def base_inv(v):
    """已有发票号值规范化：分排期惯例是 发票号-月.日(如 88718-12.26)，还原为纯发票号用于比对"""
    return DATE_SUFFIX_RE.sub('', str(v).strip())

# 总排期固定列（1-based）
TOTAL_COLS = {"country": 3, "po": 4, "cust": 5, "poline": 6, "hh": 7, "qty": 9, "date": 13, "inv": 29}

def detect_cols(ws):
    """分排期：扫描前5行识别关键列（按优先级两趟），返回 dict 或 None"""
    found = {}
    def ht(c):
        return str(c.value).replace("\n", "").replace(" ", "") if isinstance(c.value, str) else ""
    rows = list(ws.iter_rows(min_row=1, max_row=5, max_col=min(ws.max_column or 1, 90)))
    # 第一趟：高优先级精确匹配
    for row in rows:
        for c in row:
            t = ht(c)
            if not t: continue
            if "inv" not in found and re.search(r'发票|發票|invoice', t, re.I):
                found["inv"] = c.column
            elif "cust" not in found and re.search(r'客户.*PO|客PO|consignee', t, re.I):
                found["cust"] = c.column
            elif "po" not in found and re.fullmatch(r'PO号?|PO#|PONO|合同', t, re.I):
                found["po"] = c.column
            elif "poline" not in found and re.fullmatch(r'SKU(No\.)?', t, re.I):
                found["poline"] = c.column
            elif "hh" not in found and "货号" in t:
                found["hh"] = c.column
            elif "qty" not in found and "数量" in t:
                found["qty"] = c.column
            elif "date" not in found and re.search(r'出货|走货', t) and "接单" not in t:
                found["date"] = c.column
            elif "mark" not in found and re.search(r'是否.*入', t):
                found["mark"] = c.column
    # 第二趟：宽松兜底
    for row in rows:
        for c in row:
            t = ht(c)
            if not t: continue
            if "po" not in found and re.search(r'\bPO\b|合同', t, re.I) and "客户" not in t:
                found["po"] = c.column
            elif "hh" not in found and re.search(r'ITEM', t, re.I):
                found["hh"] = c.column
            elif "date" not in found and "日期" in t and not re.search(r'接单|验货|驗貨|贴纸|箱唛|订购|物料|日期码|收到|计划验', t):
                found["date"] = c.column
    return found if ("po" in found and "inv" in found) else None

def process(schedule_path, out_dir, matcher: UnifiedMatcher, log=print):
    """返回 (结果文件路径, 分类清单路径, stats)"""
    wb = openpyxl.load_workbook(schedule_path)  # 保留公式与格式
    is_total = "总排期" in wb.sheetnames
    sheets = [wb["总排期"]] if is_total else list(wb.worksheets)
    log(f"{'总排期模式' if is_total else '分排期模式'}: {len(sheets)} 个sheet")
    # 分排期惯例：发票号带录入日期后缀(如 94894-8.14)，并同步"是否已入"列；总排期保持纯数字
    today_md = f"{datetime.datetime.now().month}.{datetime.datetime.now().day}"
    fill_text = (lambda inv: str(inv)) if is_total else (lambda inv: f"{inv}-{today_md}")

    results = []   # 分类明细
    stats = Counter()
    for ws in sheets:
        cols = TOTAL_COLS if is_total else detect_cols(ws)
        if not cols:
            log(f"  sheet[{ws.title}] 未识别到PO/发票号列，跳过"); continue
        inv_col = cols["inv"]

        # 第一遍：收集全部行（含已有发票号的），按 (PO,客PO,货号候选) 分组汇总
        pending, prefilled = [], []
        group_qty = {}
        for r in range(2, (ws.max_row or 1) + 1):
            po = norm_code(ws.cell(r, cols["po"]).value)
            if not po: continue
            existing = ws.cell(r, inv_col).value
            poline = norm_code(ws.cell(r, cols["poline"]).value) if "poline" in cols else ""
            cust = norm_code(ws.cell(r, cols["cust"]).value) if "cust" in cols else ""
            hh = ws.cell(r, cols["hh"]).value if "hh" in cols else ""
            qty_v = ws.cell(r, cols["qty"]).value if "qty" in cols else None
            qty = float(qty_v) if isinstance(qty_v, (int, float)) else None
            row_date = parse_date(ws.cell(r, cols["date"]).value) if "date" in cols else None
            country = str(ws.cell(r, cols["country"]).value or "").strip() if "country" in cols else ""
            cs = code_candidates(hh or "")
            # 姊妹行分组含客PO：同一PO下不同客PO的相同货号是不同组（用户业务规则）
            gkey = (po, cust or "", tuple(sorted(cs))) if cs else (po, cust or "", ("__RAW__", str(hh or "")))
            rec = {"r": r, "po": po, "poline": poline or "", "cust": cust or "", "hh": str(hh or ""),
                   "qty": qty, "date": row_date, "country": country, "gkey": gkey, "cs": cs}
            if existing is not None and str(existing).strip():
                stats["已有发票号(跳过)"] += 1
                rec["pre"] = str(existing).strip()
                rec["pre_base"] = base_inv(existing)
                prefilled.append(rec)
            else:
                # 分组总量只统计待填行（已填的行已有归属，不参与分批合并与数量护栏的口径）
                if qty is not None:
                    group_qty[gkey] = group_qty.get(gkey, 0.0) + qty
                pending.append(rec)

        # 第二遍：逐行分类并写入
        filled = 0
        unfilled = []
        for x in pending:
            res = matcher.classify(x["po"], x["poline"], x["hh"], x["qty"], x["date"], x["country"],
                                   sibling_qty=group_qty.get(x["gkey"]), cust_po=x["cust"])
            inv_cell = ws.cell(x["r"], inv_col)
            if res["class"] == "fill":
                inv_cell.value = fill_text(res["inv"])
                inv_cell.fill = PINK_FILL
                if "mark" in cols:
                    ws.cell(x["r"], cols["mark"]).value = "已入"
                filled += 1
                stats["新填入"] += 1
                x["final"] = res["inv"]
            else:
                if res["class"] == "review":
                    stats["需人工复核"] += 1
                    cands = res.get("candidates")
                    if cands:  # 不确定但有候选：候选写进格子并标绿，人工定夺后直接改
                        inv_cell.value = str(cands)
                        inv_cell.fill = GREEN_FILL
                else:
                    stats["无发票记录"] += 1
                x["cls"] = res["class"]
                unfilled.append((x, res))
            results.append({"sheet": ws.title, "row": x["r"], "po": x["po"], "poline": x["poline"], "cust": x["cust"],
                            "hh": x["hh"], "qty": x["qty"], "date": x["date"].isoformat() if x["date"] else "",
                            "country": x["country"], "class": res["class"],
                            "inv": res.get("inv") or res.get("candidates") or "",
                            "source": res.get("source", ""), "note": res.get("note", "")})
        log(f"  sheet[{ws.title}] 新填入 {filled} 行")

        # 第三遍：整票闭合补齐（2025发票库）
        # 同组（PO+客PO+货号）只有唯一一张发票含该货号，且 该票该货号数量 == 组内[已填该票的行+空行]数量合计
        # → 空行补上该票，数量账完全闭合，文员无需再查
        groups = {}
        for x in pending:
            groups.setdefault(x["gkey"], {"rows": []})["rows"].append(x)
        for x in prefilled:
            groups.setdefault(x["gkey"], {"rows": []})["rows"].append(x)
        closed = 0
        for gkey, g in groups.items():
            rows = g["rows"]
            empty = [x for x in rows if "final" not in x and "pre" not in x]
            if not empty: continue
            po, cs_key = gkey[0], gkey[2]
            if cs_key and cs_key[0] == "__RAW__": continue
            cs = set(cs_key)
            invs = [i for i in matcher.po2inv.get(po, [])
                    if cs & set(matcher.inv_items.get(i, {}))]
            if len(invs) != 1: continue
            inv = invs[0]
            if any(not isinstance(x["qty"], float) for x in rows): continue
            inv_qty = sum(q for c in cs if c in matcher.inv_items.get(inv, {}) for q in matcher.inv_items[inv][c])
            acc = sum(x["qty"] for x in rows if x.get("pre_base") == inv or x.get("final") == inv)
            acc += sum(x["qty"] for x in empty)
            if abs(acc - inv_qty) > 1e-6: continue
            if any(x.get("final") and x["final"] != inv for x in rows): continue
            for x in empty:
                cell = ws.cell(x["r"], inv_col)
                cell.value = fill_text(inv)
                cell.fill = PINK_FILL
                if "mark" in cols:
                    ws.cell(x["r"], cols["mark"]).value = "已入"
                x["final"] = inv
                x["kind"] = "closure"
                closed += 1
                # 回填统计：该行在第二遍已计入 复核/无发票，闭合后改计
                if x.get("cls") == "review":
                    stats["需人工复核"] -= 1
                elif x.get("cls"):
                    stats["无发票记录"] -= 1
                for y in results:
                    if y["sheet"] == ws.title and y["row"] == x["r"]:
                        y.update({"class": "fill", "inv": inv, "source": "2025发票原件",
                                  "note": f"整票闭合补齐:{inv}该货号数量{inv_qty:.0f}=组内行合计"})
        if closed:
            stats["整票闭合补齐"] = stats.get("整票闭合补齐", 0) + closed
            log(f"  sheet[{ws.title}] 整票闭合补齐 {closed} 行")

        # 第四遍（兜底）：供给护栏（跨组按发票货号数量核算，含闭合补齐的行）
        # 需求=所有填该票的行(含原有+新填+闭合)按货号合计，供给=发票该货号数量
        # 1:1数量口径且需求>供给 → 本次填的相关行全部翻转为复核（无法确定哪行该得，宁可人工）
        demand = {}
        def inv_lines_of(inv):
            items = None
            if inv in matcher.db: items = matcher.db[inv]["items"]
            elif inv in matcher.zb_inv_items: items = matcher.zb_inv_items[inv]
            if not items: return None
            cands = set()
            for it in items: cands |= code_candidates(it["code"])
            out = {}
            for c in cands:
                out[c] = line_qtys_for(items, {c})
            return out
        for x in pending:
            v = x.get("final")
            if not v or "/" in str(v) or x["qty"] is None: continue
            lines = inv_lines_of(v)
            if not lines: continue
            for c in x["cs"]:
                if c not in lines: continue
                d = demand.setdefault((v, c), {"lines": lines[c], "ours": [], "pre_qty": 0.0})
                d["ours"].append(x)
        for x in prefilled:
            v = x.get("pre_base")
            if not v or "/" in str(v) or x["qty"] is None: continue
            lines = inv_lines_of(v)
            if not lines: continue
            for c in x["cs"]:
                if c not in lines: continue
                d = demand.setdefault((v, c), {"lines": lines[c], "ours": [], "pre_qty": 0.0})
                d["pre_qty"] += x["qty"]
        flipped = 0
        for (inv, c), d in demand.items():
            if not d["ours"]: continue
            supply = sum(d["lines"])
            line_set = set(d["lines"])
            all_rows = d["ours"]
            total = sum(x["qty"] for x in all_rows) + d["pre_qty"]
            one2one = all(x["qty"] in line_set for x in all_rows)
            if one2one and total > supply + 1e-6:
                for x in all_rows:
                    cell = ws.cell(x["r"], inv_col)
                    cell.value = str(inv)  # 绿色显示争议候选
                    cell.fill = GREEN_FILL
                    x.pop("final", None)
                    x["cls"] = "review"
                    flipped += 1
                    if x.get("kind") == "closure":
                        stats["整票闭合补齐"] -= 1
                    else:
                        stats["新填入"] -= 1
                    stats["需人工复核"] += 1
                    for y in results:
                        if y["sheet"] == ws.title and y["row"] == x["r"]:
                            y.update({"class": "review", "inv": inv, "source": y.get("source", ""),
                                      "note": f"供给护栏:{inv}货号{c}仅{supply:.0f}，全部需求{total:.0f}不够分，转人工"})
        if flipped:
            log(f"  sheet[{ws.title}] 供给护栏翻转 {flipped} 行为复核")
            stats["供给护栏转复核"] = stats.get("供给护栏转复核", 0) + flipped

    base = os.path.splitext(os.path.basename(schedule_path))[0]
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(out_dir, f"{base}_已入发票_{ts}.xlsx")
    wb.save(out_path)
    wb.close()

    # 分类清单
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    cwb = Workbook()
    hdr_fill = PatternFill("solid", fgColor="4472C4"); hdr_font = Font(bold=True, color="FFFFFF")
    ws0 = cwb.active; ws0.title = "汇总"
    ws0.append(["分类", "行数"])
    for k, v in stats.most_common(): ws0.append([k, v])
    ws0.column_dimensions["A"].width = 24; ws0.column_dimensions["B"].width = 12
    headers = ["sheet", "行号", "PO", "PO行号", "客PO", "货号#", "数量", "走货期", "国家", "分类", "发票号/候选", "来源", "说明"]
    for cls, title in [("fill", "新填入"), ("review", "需人工复核"), ("none", "无发票记录")]:
        w = cwb.create_sheet(title)
        w.append(headers)
        for c in w[1]: c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center")
        for x in results:
            if x["class"] == cls:
                w.append([x["sheet"], x["row"], x["po"], x["poline"], x["cust"], x["hh"], x["qty"], x["date"],
                          x["country"], x["class"], x["inv"], x["source"], x["note"]])
        for i, wd in enumerate([16, 8, 14, 16, 14, 22, 10, 12, 10, 8, 26, 14, 40], 1):
            w.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = wd
        w.freeze_panes = "A2"
    cls_path = os.path.join(out_dir, f"{base}_分类清单_{ts}.xlsx")
    cwb.save(cls_path); cwb.close()
    return out_path, cls_path, dict(stats)

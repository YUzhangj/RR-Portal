#!/usr/bin/env python3
"""从上传的发票原件 Excel 构建发票库 JSON
- kind=2025: sheet-per-invoice 格式（如 2025 ZURU（最新）1-6月/7-12月）
- kind=zb:   2026 ZB 减价格式（汇总表 sheet + 实体发票 sheet）
"""
import openpyxl, re, json, datetime, os
from collections import defaultdict

PO_RE = re.compile(r'PO\s*NO\s*[:：]?\s*([A-Za-z0-9][A-Za-z0-9-]{3,})', re.I)
PO_CELL_RE = re.compile(r'45\d{8}(-\d+)?')

def _num(v):
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str):
        try: return float(v.replace(',', '').strip())
        except ValueError: return None
    return None

def _extract_2025_raw(path, log=print):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    raw = {}
    for name in wb.sheetnames:
        ws = wb[name]
        pos, items, inv_date = set(), [], None
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 60, 90), max_col=13):
            for c in row:
                v = c.value
                if v is None: continue
                if isinstance(v, str):
                    m = PO_RE.search(v)
                    if m: pos.add(m.group(1))
                elif isinstance(v, datetime.datetime) and inv_date is None and c.row <= 12 and 5 <= c.column <= 9:
                    inv_date = v.date().isoformat()
            a = row[0].value if len(row) > 0 else None
            q = _num(row[3].value) if len(row) > 3 else None
            e = row[4].value if len(row) > 4 else None
            if a is not None and q is not None and isinstance(e, str) and e.strip().upper() == 'PCS':
                code = a.strip() if isinstance(a, str) else str(a)
                if code.startswith('MA#'): continue
                price = _num(row[6].value) if len(row) > 6 else None   # G列单价
                amount = _num(row[8].value) if len(row) > 8 else None  # I列金额
                items.append({"code": code, "qty": q, "price": price, "amount": amount})
        raw[name] = {"pos": sorted(pos), "items": items, "date": inv_date}
    wb.close()
    log(f"  {os.path.basename(path)}: {len(raw)} 个sheet")
    return raw

def build_2025(paths, data_dir, log=print):
    raw = {}
    for p in paths:
        raw.update(_extract_2025_raw(p, log))
    def merge_key(name):
        """返回 (归入的发票号, 是否重复页)。
        尾点sheet(如 94907.)：与基名sheet明细完全一致视为重复页(不加明细)，
        无基名sheet则规范化名称(去掉尾点)；明细不同则不合并，防数量翻倍。"""
        if name.endswith('.'):
            b = name[:-1]
            if b in raw:
                si = sorted((i["code"], i["qty"]) for i in raw[name]["items"])
                sb = sorted((i["code"], i["qty"]) for i in raw[b]["items"])
                if si == sb and raw[name]["pos"] == raw[b]["pos"]:
                    return b, True
                return name, False
            return b, False
        m = re.match(r'^(\d+(?:\.\d+)?)-(\d+)$', name)
        if m and m.group(1) in raw: return m.group(1), False
        return name, False

    db = {}
    for name, d in raw.items():
        b, is_dup = merge_key(name)
        if b not in db:
            db[b] = {"pos": set(), "items": [], "date": d["date"], "pages": []}
        db[b]["pos"].update(d["pos"])
        if not is_dup:
            db[b]["items"].extend(d["items"])
        db[b]["pages"].append(name)
        if not db[b]["date"] and d["date"]: db[b]["date"] = d["date"]
    out = {k: {"pos": sorted(v["pos"]), "items": v["items"], "date": v["date"], "pages": v["pages"]} for k, v in db.items()}
    po2inv = defaultdict(list)
    for inv, d in out.items():
        for po in d["pos"]: po2inv[po].append(inv)
    json.dump(out, open(os.path.join(data_dir, "发票DB_2025全年.json"), "w"), ensure_ascii=False)
    json.dump(dict(po2inv), open(os.path.join(data_dir, "PO到发票_2025全年.json"), "w"), ensure_ascii=False)
    log(f"2025发票库: {len(out)} 张发票 / {len(po2inv)} 个PO")
    return len(out), len(po2inv)

def detect_file_type(path):
    """自动识别发票文件类型：'zb' = 含累计汇总表的首张sheet；'2025' = sheet-per-invoice 格式"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    first = wb[names[0]]
    is_zb = False
    if "汇总" in names[0]:
        is_zb = True
    else:
        for row in first.iter_rows(min_row=1, max_row=8, max_col=4, values_only=True):
            vals = [str(v) for v in row if v is not None]
            if any("汇总" in v for v in vals) or any(v.strip().upper().startswith("ZB") for v in vals):
                is_zb = True
                break
    if not is_zb:
        # 2025格式：sheet名基本是发票号
        import re as _re
        inv_like = sum(1 for n in names[:30] if _re.match(r'^\d{4,6}([-.]\d+)?$', n))
        if inv_like < max(3, len(names[:30]) // 3):
            wb.close()
            raise ValueError(f"无法识别的发票文件格式: {os.path.basename(path)}")
    wb.close()
    return "zb" if is_zb else "2025"

def build_zb(paths, data_dir, log=print):
    """按文件修改时间新→旧处理，setdefault保留最新；实体sheet按名字去重（同名以新文件为准）"""
    summary, sheet2po, detail = {}, {}, []
    seen_sheets = set()
    for p in sorted(paths, key=lambda x: -os.path.getmtime(x)):
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        # 汇总表（首张sheet）
        ws = wb[wb.sheetnames[0]]
        n = 0
        for row in ws.iter_rows(min_row=1, max_col=9, values_only=True):
            inv = str(row[0]).strip() if row[0] else ""
            if not inv.upper().startswith("ZB"): continue
            dt = str(row[1])[:10] if row[1] else None
            poline = str(row[2]).strip() if row[2] else ""
            if poline:
                summary.setdefault(poline, {"inv": inv, "date": dt}); n += 1
        log(f"  {os.path.basename(p)} 汇总表: {n} 条")
        # 实体发票 sheet：提取 PO行号(J8区)、客PO(Ultimate consignee PO#列)、货号明细(C列+H列数量)
        ns = 0
        for name in wb.sheetnames[1:]:
            if name in seen_sheets:
                continue  # 同名sheet以更新的文件为准
            seen_sheets.add(name)
            ws2 = wb[name]
            po = None
            for r in ws2.iter_rows(min_row=6, max_row=10, max_col=10):
                for c in r:
                    if isinstance(c.value, str) and PO_CELL_RE.fullmatch(c.value.strip()):
                        po = c.value.strip()
            cust, items = [], []
            hdr_row = None
            for r in ws2.iter_rows(min_row=11, max_row=min(ws2.max_row or 55, 60), max_col=13):
                b = r[1].value if len(r) > 1 else None
                if hdr_row is None and isinstance(b, str) and "onsignee" in b and "PO" in b:
                    hdr_row = r[0].row
                    continue
                if hdr_row and isinstance(r[0].value, (int, float)) and b:
                    cust.append(str(b).strip())
                # 货号行：C列货号(非MA#) + H列数量，排除 Final Price 汇总行
                cc = r[2].value if len(r) > 2 else None
                qq = _num(r[7].value) if len(r) > 7 else None
                ee = r[4].value if len(r) > 4 else None
                if cc is not None and qq is not None:
                    code = cc.strip() if isinstance(cc, str) else str(cc)
                    if code.startswith('MA#'): continue
                    if isinstance(ee, str) and 'Final Price' in ee: continue
                    price = _num(r[8].value) if len(r) > 8 else None    # I列减价后单价
                    amount = _num(r[9].value) if len(r) > 9 else None   # J列金额
                    items.append({"code": code, "qty": qq, "price": price, "amount": amount})
            if po:
                sheet2po.setdefault(po, []).append(name); ns += 1
            detail.append({"sheet": name, "po": po, "cust": cust, "items": items})
        log(f"  {os.path.basename(p)} 实体sheet: {ns} 张")
        wb.close()
    json.dump(summary, open(os.path.join(data_dir, "ZB映射_PO行号到发票.json"), "w"), ensure_ascii=False)
    json.dump(sheet2po, open(os.path.join(data_dir, "ZB实体sheet_PO映射.json"), "w"), ensure_ascii=False)
    json.dump(detail, open(os.path.join(data_dir, "ZB实体sheet_明细.json"), "w"), ensure_ascii=False)
    log(f"ZB发票库: 汇总表 {len(summary)} 条 / 实体sheet {len(sheet2po)} 个PO / 明细 {len(detail)} 张")
    return len(summary), len(sheet2po)
